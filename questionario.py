import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('questionario apn.xlsx')
questionnaire = wb.active
pergunta = []
resposta = []
for row in range(2, questionnaire.max_row + 1):
    p = questionnaire.cell(row, 2)
    pergunta.append(p.value)
i = 0
while i < 80:
    print(pergunta[i])
    r = input('> ')
    if r.lower() == 's':
        resposta.append(i+1)
        i += 1
    elif r.lower() == 'n':
        i += 1
    else:
        i = i
ativo = [3, 5, 7, 9, 13, 20, 26, 27, 35, 37, 41, 43, 46, 48, 51, 61, 67, 74, 75, 77]
reflexivo = [10, 16, 18, 19, 28, 31, 32, 34, 36, 39, 42, 44, 49, 55, 58, 63, 65, 69, 70, 79]
teorico = [2, 4, 6, 11, 15, 17, 21, 23, 25, 29, 33, 45, 50, 54, 60, 64, 66, 71, 78, 80]
pragmatico = [1, 8, 12, 14, 22, 24, 30, 38, 40, 47, 52, 53, 56, 57, 59, 62, 68, 72, 73, 76]

result_pragmatico = list(set(pragmatico) & set(resposta))
result_ativo = list(set(ativo) & set(resposta))
result_teorico = list(set(teorico) & set(resposta))
result_reflexivo = list(set(reflexivo) & set(resposta))

result_title = questionnaire.cell(1, 3)
result_title.value = 'Resultados'

ativo_title = questionnaire.cell(2, 3)
ativo_title.value = 'Ativo'
ativo_valor = questionnaire.cell(2, 4)
ativo_valor.value = len(result_ativo)

teorico_title = questionnaire.cell(3, 3)
teorico_title.value = 'Teórico'
teorico_valor = questionnaire.cell(3, 4)
teorico_valor.value = len(result_teorico)

pragmatico_title = questionnaire.cell(4, 3)
pragmatico_title.value = 'Pragmático'
pragmatico_valor = questionnaire.cell(4, 4)
pragmatico_valor.value = len(result_pragmatico)

reflexivo_title = questionnaire.cell(5, 3)
reflexivo_title.value = 'Reflexivo'
reflexivo_valor = questionnaire.cell(5, 4)
reflexivo_valor.value = len(result_reflexivo)

chart = BarChart()
data = Reference(questionnaire,
                 min_row=2,
                 max_row=5,
                 min_col=3,
                 max_col=4)

chart.add_data(data)
questionnaire.add_chart(chart, "F2")

wb.save("Resultado_Questionario_APN.xlsx")
